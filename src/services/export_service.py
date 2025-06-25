"""Export utilities for generating monthly reports."""

from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import logging
from typing import Callable, Iterable
import uuid

import pandas as pd
import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib import font_manager

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference

from reportlab.pdfgen.canvas import Canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader

from ..models import FuelEntry
from .storage_service import StorageService

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting monthly reports."""

    def __init__(self, storage: StorageService) -> None:
        self.storage = storage
        self._tmpdirs: list[TemporaryDirectory[str]] = []

    def cleanup(self) -> None:
        """Remove any temporary directories created during exports."""
        for tmp in self._tmpdirs:
            tmp.cleanup()
        self._tmpdirs.clear()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def export_monthly_pdf(self, month: str, vehicle_id: int | None) -> Path:
        """Create a detailed monthly PDF report and return its path."""
        year, mon = (int(x) for x in month.split("-"))
        entries = self.storage.list_entries_for_month(year, mon, vehicle_id)
        df = self._entries_to_df(entries)
        tmp = TemporaryDirectory()
        self._tmpdirs.append(tmp)
        out_path = Path(tmp.name) / f"report_{month}.pdf"

        charts: list[Path] = []

        def p1(c: Canvas, f: str) -> None:
            charts.extend(
                self._page_summary(c, f, df, month, vehicle_id, out_path.parent)
            )

        def p2(c: Canvas, f: str) -> None:
            charts.extend(self._page_weekly(c, f, df, out_path.parent))

        def p3(c: Canvas, f: str) -> None:
            charts.extend(self._page_compare(c, f, df, out_path.parent))

        try:
            self._build_pdf(out_path, [p1, p2, p3])
        finally:
            for ch in charts:
                if ch.exists():
                    ch.unlink()
        return out_path

    def export_monthly_xlsx(self, month: str) -> Path:
        """Create a detailed monthly Excel report and return its path."""
        year, mon = (int(x) for x in month.split("-"))
        entries = self.storage.list_entries_for_month(year, mon)
        df = self._entries_to_df(entries)
        tmp = TemporaryDirectory()
        self._tmpdirs.append(tmp)
        out_path = Path(tmp.name) / f"report_{month}.xlsx"
        self._write_excel(out_path, df)
        return out_path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    def _entries_to_df(self, entries: list[FuelEntry]) -> pd.DataFrame:
        data = []
        for e in entries:
            dist = None
            if e.odo_after is not None:
                dist = e.odo_after - e.odo_before
            data.append(
                {
                    "date": e.entry_date,
                    "vehicle_id": e.vehicle_id,
                    "fuel_type": e.fuel_type,
                    "odo_before": e.odo_before,
                    "odo_after": e.odo_after,
                    "distance": dist,
                    "liters": e.liters,
                    "amount_spent": e.amount_spent,
                }
            )
        df = pd.DataFrame(data)
        if not df.empty:
            df.sort_values("date", inplace=True)
        return df

    def _get_font(self) -> str:
        try:
            font_path = font_manager.findfont(
                "Noto Sans Thai", fallback_to_default=False
            )  # type: ignore[operator]
        except Exception as exc:  # pragma: no cover - not found
            logger.debug("ไม่พบ Noto Sans Thai: %s", exc)
            font_path = ""

        if font_path and Path(font_path).exists():
            try:
                pdfmetrics.registerFont(
                    TTFont("NotoSansThai", font_path, subsetting=True)
                )
                return "NotoSansThai"
            except Exception as exc:  # pragma: no cover - bad font
                logger.warning("ไม่สามารถลงทะเบียน NotoSansThai ได้: %s", exc)

        try:
            fallback_path = font_manager.findfont("Tahoma")  # type: ignore[operator]
            pdfmetrics.registerFont(TTFont("Tahoma", fallback_path, subsetting=True))
            return "Tahoma"
        except Exception as exc:  # pragma: no cover - fallback issue
            logger.warning("ไม่สามารถลงทะเบียนฟอนต์สำรองได้: %s", exc)
            return "Helvetica"

    def _plot_dual_axis(self, fig: Figure, df: pd.DataFrame) -> None:
        """Plot liters and km/L on dual axes."""
        ax1 = fig.add_subplot(111)
        ax2: Axes = ax1.twinx()

        if df.empty:
            return

        daily = df.groupby("date")[["distance", "liters"]].sum()
        daily["km_per_l"] = daily["distance"] / daily["liters"]

        ax1.bar(daily.index, daily["liters"], color="tab:blue", alpha=0.6, label="L")
        ax2.plot(
            daily.index,
            daily["km_per_l"],
            color="tab:red",
            marker="o",
            label="km/L",
        )
        ax1.set_ylabel("Liters")
        ax2.set_ylabel("km/L")
        ax1.set_xlabel("Date")
        fig.autofmt_xdate()
        ax1.legend(loc="upper left")
        ax2.legend(loc="upper right")
        fig.tight_layout()

    def _build_pdf(
        self, path: Path, pages: Iterable[Callable[[Canvas, str], None]]
    ) -> None:
        """Render each page builder to the PDF."""
        font_name = self._get_font()
        canvas = Canvas(str(path), pagesize=A4)
        page_list = list(pages)
        for i, builder in enumerate(page_list):
            builder(canvas, font_name)
            if i < len(page_list) - 1:
                canvas.showPage()
        canvas.save()

    # ------------------------------------------------------------------
    # Page builders
    # ------------------------------------------------------------------
    def _page_summary(
        self,
        canvas: Canvas,
        font: str,
        df: pd.DataFrame,
        month: str,
        vehicle_id: int | None,
        tmp_dir: Path,
    ) -> list[Path]:
        chart = tmp_dir / f"{uuid.uuid4().hex}.png"
        fig = plt.Figure(figsize=(6, 4))
        self._plot_dual_axis(fig, df)
        fig.savefig(chart, format="png")
        plt.close(fig)

        total_distance = float(df["distance"].fillna(0).sum())
        total_liters = float(df["liters"].fillna(0).sum())
        km_per_l = total_distance / total_liters if total_liters else 0.0

        canvas.setFont(font, 16)
        title = f"รายงานประจำเดือน {month}"
        if vehicle_id is not None:
            vehicle = self.storage.get_vehicle(vehicle_id)
            if vehicle is not None:
                title += f" ยานพาหนะ {vehicle.name}"
            else:
                title += f" ยานพาหนะ {vehicle_id}"
        canvas.drawString(40, 800, title)

        canvas.setFont(font, 12)
        canvas.drawString(40, 780, f"ระยะทางรวม: {total_distance:.1f} km")
        canvas.drawString(40, 764, f"ปริมาณเชื้อเพลิง: {total_liters:.1f} L")
        canvas.drawString(40, 748, f"อัตราสิ้นเปลือง: {km_per_l:.2f} km/L")

        canvas.drawImage(
            ImageReader(str(chart)),
            40,
            470,
            width=500,
            preserveAspectRatio=True,
        )
        return [chart]

    def _weekly_pivot(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            cols = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
            return pd.DataFrame(columns=cols, index=pd.Index([], name="iso_week"))

        temp = df.copy()
        temp["weekday"] = pd.to_datetime(temp["date"]).dt.day_name().str[:3]
        temp["iso_week"] = temp["date"].apply(
            lambda d: f"{d.isocalendar().year}-W{d.isocalendar().week:02d}"
        )
        pivot = temp.pivot_table(
            index="iso_week", columns="weekday", values="liters", aggfunc="sum"
        ).fillna(0)
        cols = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for col in cols:
            if col not in pivot.columns:
                pivot[col] = 0
        pivot = pivot[cols]
        pivot.sort_index(inplace=True)
        return pivot

    def _page_weekly(
        self,
        canvas: Canvas,
        font: str,
        df: pd.DataFrame,
        tmp_dir: Path,
    ) -> list[Path]:
        pivot = self._weekly_pivot(df)
        canvas.setFont(font, 16)
        canvas.drawString(40, 800, "สรุปรายสัปดาห์")

        canvas.setFont(font, 10)
        y = 780
        headers = ["Week"] + pivot.columns.tolist()
        for i, h in enumerate(headers):
            canvas.drawString(40 + i * 70, y, str(h))
        y -= 16
        for week, row in pivot.iterrows():
            canvas.drawString(40, y, str(week))
            for i, col in enumerate(pivot.columns):
                canvas.drawString(40 + (i + 1) * 70, y, f"{row[col]:.1f}")
            y -= 16
            if y < 200:
                break

        chart = tmp_dir / f"{uuid.uuid4().hex}.png"
        fig = plt.Figure(figsize=(6, 4))
        ax = fig.add_subplot(111)
        daily = df.groupby("date")["liters"].sum()
        if not daily.empty:
            ax.bar(daily.index, daily.values, color="tab:blue")
        ax.set_ylabel("Liters")
        ax.set_xlabel("Date")
        fig.autofmt_xdate()
        fig.tight_layout()
        fig.savefig(chart, format="png")
        plt.close(fig)

        canvas.drawImage(
            ImageReader(str(chart)),
            40,
            420,
            width=500,
            preserveAspectRatio=True,
        )
        return [chart]

    def _page_compare(
        self,
        canvas: Canvas,
        font: str,
        df: pd.DataFrame,
        tmp_dir: Path,
    ) -> list[Path]:
        canvas.setFont(font, 16)
        canvas.drawString(40, 800, "เปรียบเทียบยานพาหนะ")
        if df.empty:
            return []

        data = df.groupby("vehicle_id")[["distance", "liters"]].sum()
        data["km_per_l"] = data["distance"] / data["liters"]
        data.sort_values("km_per_l", ascending=False, inplace=True)

        names = {v.id: v.name for v in self.storage.list_vehicles()}
        labels = [names.get(idx, str(idx)) for idx in data.index]

        chart = tmp_dir / f"{uuid.uuid4().hex}.png"
        fig = plt.Figure(figsize=(6, 4))
        ax1 = fig.add_subplot(111)
        ax2 = ax1.twinx()
        ax1.bar(
            labels,
            data["liters"],
            color="tab:blue",
            alpha=0.6,
            label="L",
        )
        ax2.plot(
            labels,
            data["km_per_l"],
            color="tab:red",
            marker="o",
            label="km/L",
        )
        ax1.set_ylabel("Liters")
        ax2.set_ylabel("km/L")
        ax1.set_xlabel("Vehicle")
        ax1.legend(loc="upper left")
        ax2.legend(loc="upper right")
        fig.tight_layout()
        fig.savefig(chart, format="png")
        plt.close(fig)

        canvas.drawImage(
            ImageReader(str(chart)),
            40,
            460,
            width=500,
            preserveAspectRatio=True,
        )
        return [chart]

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def _write_excel(self, path: Path, df: pd.DataFrame) -> None:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Entries"

        headers = [
            "date",
            "fuel_type",
            "odo_before",
            "odo_after",
            "distance",
            "liters",
            "amount_spent",
        ]
        ws_data.append(headers)
        for r in df.itertuples(index=False):
            ws_data.append(
                [
                    r.date,
                    r.fuel_type,
                    r.odo_before,
                    r.odo_after,
                    r.distance,
                    r.liters,
                    r.amount_spent,
                ]
            )

        ws_month = wb.create_sheet("Summary")

        total_distance = float(df["distance"].fillna(0).sum())
        total_liters = float(df["liters"].fillna(0).sum())
        km_per_l = total_distance / total_liters if total_liters else 0.0

        ws_month.append(["metric", "value"])
        ws_month.append(["distance", total_distance])
        ws_month.append(["liters", total_liters])
        ws_month.append(["km_per_l", km_per_l])

        daily = df.groupby("date")[["distance", "liters"]].sum().reset_index()
        daily["km_per_l"] = daily["distance"] / daily["liters"]

        ws_month.append([])
        ws_month.append(["date", "liters", "km_per_l"])
        for r in daily.itertuples(index=False):
            ws_month.append([r.date, r.liters, r.km_per_l])

        start = ws_month.max_row - len(daily) + 1
        chart = BarChart()
        data = Reference(
            ws_month, min_col=2, min_row=start, max_row=start + len(daily) - 1
        )
        cats = Reference(
            ws_month, min_col=1, min_row=start, max_row=start + len(daily) - 1
        )
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(cats)
        chart.y_axis.title = "Liters"

        line = LineChart()
        line.add_data(
            Reference(
                ws_month, min_col=3, min_row=start, max_row=start + len(daily) - 1
            ),
            titles_from_data=False,
        )
        line.y_axis.axId = 200
        _ = line.y_axis.axId  # reference to avoid vulture warning
        chart += line
        ws_month.add_chart(chart, "E2")

        ws_week = wb.create_sheet("Weekly")
        pivot = self._weekly_pivot(df)
        ws_week.append(["iso_week"] + pivot.columns.tolist())
        for idx, row in pivot.iterrows():
            ws_week.append([idx] + [float(row[c]) for c in pivot.columns])

        wb.save(path)
